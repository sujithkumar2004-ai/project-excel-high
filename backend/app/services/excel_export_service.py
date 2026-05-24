from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.config import settings
from app.models import ExtractedRecordRow

EXCEL_COLUMNS = [
    ("S.No", "row_number"),
    ("Hospital Registration No", "hospital_registration_no"),
    ("Name of Patient", "patient_name"),
    ("Age/Sex", "age"),
    ("Provisional Diagnosis", "provisional_diagnosis"),
    ("Name of Procedure", "procedure_name"),
    ("Final Diagnosis", "final_diagnosis"),
    ("Name of Surgeon / Anesthetist / Staff", "surgeon_name"),
    ("OT No", "ot_number"),
    ("Date", "procedure_date"),
    ("Start Time", "start_time"),
    ("End Time", "end_time"),
    ("Type of Anesthesia", "anesthesia_type"),
]


def generate_records_excel(batch_id: int, rows: Iterable[ExtractedRecordRow]) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hospital Records"

    for column_index, (header, _) in enumerate(EXCEL_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)

    for row_index, record in enumerate(rows, start=2):
        for column_index, (_, field_name) in enumerate(EXCEL_COLUMNS, start=1):
            value = row_index - 1 if field_name == "row_number" else getattr(record, field_name, "")
            sheet.cell(row=row_index, column=column_index, value=value or "")

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)

    output_dir = Path(settings.upload_dir) / "excel"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"record_batch_{batch_id}.xlsx"
    workbook.save(output_path)
    return str(output_path)
