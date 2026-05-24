from __future__ import annotations

from datetime import datetime
from pathlib import Path
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings


def generate_excel(columns: list[str], rows: list[dict[str, str]]) -> str:
    if not columns:
        raise HTTPException(status_code=400, detail="At least one column is required for export")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reviewed Register"

    for column_index, header in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=column_index, value=str(row.get(header, "") or ""))

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 52)

    output_dir = Path(settings.export_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"reviewed_register_{uuid4().hex}.xlsx"
    workbook.save(output_path)
    return str(output_path)


def generate_ot_register_excel(columns: list[str], rows: list[dict[str, dict]], record_dir: str | Path) -> str:
    if not columns:
        raise HTTPException(status_code=400, detail="At least one column is required for export")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OT Register"
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    review_fill = PatternFill("solid", fgColor="FFF2CC")

    for column_index, header in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(columns, start=1):
            cell_payload = row.get(header) or {}
            cell = sheet.cell(row=row_index, column=column_index, value=str(cell_payload.get("value") or ""))
            if cell_payload.get("uncertain") or cell_payload.get("edited"):
                cell.fill = review_fill

    for column_index in range(1, len(columns) + 1):
        column_letter = get_column_letter(column_index)
        max_length = max(len(str(cell.value or "")) for cell in sheet[column_letter])
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

    output_dir = Path(record_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"OT_Register_{datetime.now():%Y_%m_%d}.xlsx"
    workbook.save(output_path)
    return str(output_path)
