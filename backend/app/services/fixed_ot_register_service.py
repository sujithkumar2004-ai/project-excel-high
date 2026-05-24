from __future__ import annotations

import json
import re
from functools import lru_cache
from pathlib import Path
from statistics import mean
from uuid import uuid4

import cv2
import numpy as np
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps

from app.core.config import settings

DOCUMENT_TYPE = "hospital_ot_register"

COLUMN_KEYS = [
    "s_no",
    "ipd_hospital_registration_no",
    "patient_name",
    "age_sex",
    "provisional_diagnosis",
    "procedure_name",
    "final_diagnosis",
    "surgeon_anaesthetist_staff",
    "ot",
    "date_of_procedure",
    "surgery_time",
    "end_time",
    "type_of_anaesthesia",
]

EXCEL_HEADERS = [
    "S.No",
    "IPD / Hospital Registration No.",
    "Name of the Patient",
    "Age / Sex",
    "Provisional Diagnosis",
    "Name of Procedure",
    "Final Diagnosis",
    "Name of Surgeon and Anaesthetist Staff",
    "OT",
    "Date of Procedure",
    "Surgery Time",
    "End Time",
    "Type of Anaesthesia",
    "Confidence",
    "Uncertain Fields",
]

ALLOWED_IMAGE_TYPES = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
}

MIN_CONFIDENCE = 0.52
UNCERTAIN_CONFIDENCE = 0.78
COLUMN_BOUNDARY_RATIOS = [0.065, 0.117, 0.205, 0.307, 0.375, 0.449, 0.568, 0.649, 0.728, 0.802, 0.85, 0.909, 0.955, 1.0]


async def create_fixed_ot_record(file: UploadFile) -> dict:
    record_id = uuid4().hex
    record_dir = _record_dir(record_id)
    record_dir.mkdir(parents=True, exist_ok=True)

    image_path = await _save_original_image(file, record_dir)
    processed_path = _preprocess_image(image_path, record_dir)
    extracted = _extract_table_with_paddleocr(processed_path, record_dir)
    normalized = _normalize_payload(extracted)
    _write_json(record_dir / "extracted.json", normalized)

    return _record_response(record_id, normalized, image_path, processed_path)


def get_fixed_ot_record(record_id: str) -> dict:
    record_dir = _existing_record_dir(record_id)
    data = _read_json(_active_json_path(record_dir))
    return _record_response(
        record_id,
        data,
        _first_existing(record_dir, "original.*"),
        _first_existing(record_dir, "processed.*"),
    )


def save_fixed_ot_data(record_id: str, payload: dict) -> dict:
    record_dir = _existing_record_dir(record_id)
    normalized = _normalize_payload(payload)
    _write_json(record_dir / "reviewed.json", normalized)
    return _record_response(
        record_id,
        normalized,
        _first_existing(record_dir, "original.*"),
        _first_existing(record_dir, "processed.*"),
    )


def export_fixed_ot_excel(record_id: str) -> str:
    record_dir = _existing_record_dir(record_id)
    data = _read_json(_active_json_path(record_dir))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OT_Register"
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column_index, header in enumerate(EXCEL_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_index, row in enumerate(data["rows"], start=2):
        for column_index, key in enumerate(COLUMN_KEYS, start=1):
            sheet.cell(row=row_index, column=column_index, value=row.get(key))
        sheet.cell(row=row_index, column=len(COLUMN_KEYS) + 1, value=float(row.get("confidence") or 0.0))
        sheet.cell(row=row_index, column=len(COLUMN_KEYS) + 2, value=", ".join(row.get("uncertain_fields") or []))

    for column_index in range(1, len(EXCEL_HEADERS) + 1):
        column_letter = get_column_letter(column_index)
        max_length = max(len(str(cell.value or "")) for cell in sheet[column_letter])
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)

    excel_path = record_dir / "OT_Register.xlsx"
    workbook.save(excel_path)
    return str(excel_path)


async def _save_original_image(file: UploadFile, record_dir: Path) -> str:
    content_type = file.content_type or ""
    if content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {file.filename}")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded image is empty")

    max_bytes = settings.max_upload_mb * 1024 * 1024
    if len(content) > max_bytes:
        raise HTTPException(status_code=400, detail=f"File too large: {file.filename}")

    suffix = Path(file.filename or "").suffix.lower() or ALLOWED_IMAGE_TYPES[content_type]
    if suffix == ".jpeg":
        suffix = ".jpg"
    if suffix not in {".jpg", ".png", ".webp"}:
        raise HTTPException(status_code=400, detail=f"Unsupported file extension: {file.filename}")

    image_path = record_dir / f"original{suffix}"
    image_path.write_bytes(content)
    return str(image_path)


def _preprocess_image(image_path: str, record_dir: Path) -> str:
    debug_dir = _debug_dir(record_dir)
    with Image.open(image_path) as image:
        corrected = ImageOps.exif_transpose(image).convert("RGB")
        if corrected.height > corrected.width:
            corrected = corrected.rotate(90, expand=True)
        corrected.thumbnail((2600, 2600))
        normalized_path = record_dir / "processed.jpg"
        corrected.save(normalized_path, quality=96)

    image = cv2.imread(str(normalized_path))
    if image is None:
        raise HTTPException(status_code=400, detail="Unable to read uploaded image")

    image = _deskew(image)
    cv2.imwrite(str(debug_dir / "oriented.png"), image)
    lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
    l_channel, a_channel, b_channel = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    enhanced_l = clahe.apply(l_channel)
    enhanced = cv2.cvtColor(cv2.merge((enhanced_l, a_channel, b_channel)), cv2.COLOR_LAB2BGR)
    denoised = cv2.fastNlMeansDenoisingColored(enhanced, None, 5, 5, 7, 21)
    gray = cv2.cvtColor(denoised, cv2.COLOR_BGR2GRAY)
    threshold = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 35, 8)
    cv2.imwrite(str(debug_dir / "threshold.png"), threshold)
    cv2.imwrite(str(normalized_path), denoised)
    return str(normalized_path)


def _deskew(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    coordinates = np.column_stack(np.where(binary > 0))
    if coordinates.size == 0:
        return image
    angle = cv2.minAreaRect(coordinates)[-1]
    if angle < -45:
        angle = 90 + angle
    elif angle > 45:
        angle = angle - 90
    if abs(angle) < 0.4:
        return image
    height, width = image.shape[:2]
    matrix = cv2.getRotationMatrix2D((width // 2, height // 2), angle, 1.0)
    return cv2.warpAffine(image, matrix, (width, height), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)


def _extract_table_with_paddleocr(image_path: str, record_dir: Path) -> dict:
    image = cv2.imread(image_path)
    if image is None:
        raise HTTPException(status_code=400, detail="Unable to read preprocessed image")

    grid = _detect_grid(image, record_dir)
    cells_dir = _debug_dir(record_dir) / "cells"
    cells_dir.mkdir(parents=True, exist_ok=True)

    rows: list[dict] = []
    for row_index, (top, bottom) in enumerate(grid["rows"], start=1):
        row_values: dict[str, str | None] = {}
        row_confidences: list[float] = []
        uncertain_fields: list[str] = []

        for column_index, key in enumerate(COLUMN_KEYS):
            left, right = grid["columns"][column_index]
            crop = _crop_cell(image, left, top, right, bottom)
            cell_path = cells_dir / f"{row_index:03d}_{column_index + 1:02d}.png"
            cv2.imwrite(str(cell_path), crop)
            text, confidence = _ocr_cell(crop)
            row_confidences.append(confidence)

            if not text:
                row_values[key] = None
                if confidence < MIN_CONFIDENCE:
                    uncertain_fields.append(key)
            else:
                row_values[key] = text
                if confidence < UNCERTAIN_CONFIDENCE:
                    uncertain_fields.append(key)

        if all(value is None for value in row_values.values()):
            continue

        row_payload = {
            **row_values,
            "confidence": round(mean(row_confidences), 3) if row_confidences else 0.0,
            "uncertain_fields": uncertain_fields,
        }
        rows.append(_validate_row(row_payload))

    _write_json(record_dir / "grid.json", grid)
    return {
        "document_type": DOCUMENT_TYPE,
        "columns": COLUMN_KEYS,
        "rows": rows,
    }


def _detect_grid(image: np.ndarray, record_dir: Path) -> dict:
    debug_dir = _debug_dir(record_dir)
    height, width = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    binary = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 35, -8)

    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(width // 28, 30), 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(height // 14, 20)))
    horizontal = cv2.dilate(cv2.erode(binary, horizontal_kernel, iterations=1), horizontal_kernel, iterations=2)
    vertical = cv2.dilate(cv2.erode(binary, vertical_kernel, iterations=1), vertical_kernel, iterations=2)
    grid_lines = cv2.bitwise_or(horizontal, vertical)
    cv2.imwrite(str(debug_dir / "grid_lines.png"), grid_lines)

    horizontal_lines = _line_positions(horizontal, axis="horizontal", min_length=max(width * 0.35, 120))
    vertical_lines = _vertical_positions_from_hough(gray)

    vertical_lines = _select_column_lines(vertical_lines, width)

    if len(horizontal_lines) < 3:
        horizontal_lines = _fallback_row_lines(height)

    row_bands = _data_row_bands(horizontal_lines, height)
    columns = [(int(vertical_lines[index]), int(vertical_lines[index + 1])) for index in range(len(COLUMN_KEYS))]
    rows = [(int(top), int(bottom)) for top, bottom in row_bands]
    _save_cell_box_debug(image, columns, rows, debug_dir)

    return {
        "columns": columns,
        "rows": rows,
    }


def _vertical_positions_from_hough(gray: np.ndarray) -> list[int]:
    edges = cv2.Canny(cv2.GaussianBlur(gray, (3, 3), 0), 50, 150, apertureSize=3)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold=80, minLineLength=max(gray.shape[0] // 5, 80), maxLineGap=18)
    positions: list[int] = []
    if lines is None:
        return positions
    for line in lines[:, 0]:
        x1, y1, x2, y2 = [int(value) for value in line]
        dx = x2 - x1
        dy = y2 - y1
        if dx == 0:
            angle = 90
        else:
            angle = abs(np.degrees(np.arctan2(dy, dx)))
        if 74 <= angle <= 106 and abs(y2 - y1) >= max(gray.shape[0] * 0.16, 80):
            positions.append(round((x1 + x2) / 2))
    return _merge_positions(sorted(positions), tolerance=18)


def _line_positions(mask: np.ndarray, axis: str, min_length: float) -> list[int]:
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    positions: list[int] = []
    for contour in contours:
        x, y, width, height = cv2.boundingRect(contour)
        if axis == "horizontal" and width >= min_length and height <= 18:
            positions.append(y + height // 2)
        if axis == "vertical" and height >= min_length and width <= 18:
            positions.append(x + width // 2)
    return _merge_positions(sorted(positions), tolerance=12)


def _merge_positions(values: list[int], tolerance: int) -> list[int]:
    if not values:
        return []
    groups: list[list[int]] = [[values[0]]]
    for value in values[1:]:
        if value - groups[-1][-1] <= tolerance:
            groups[-1].append(value)
        else:
            groups.append([value])
    return [round(mean(group)) for group in groups]


def _select_column_lines(lines: list[int], width: int) -> list[int]:
    detected = [line for line in lines if 0 <= line <= width]
    fallback = _fallback_column_lines(width)
    selected: list[int] = []
    for index, expected in enumerate(fallback):
        if index in {0, len(fallback) - 1}:
            selected.append(expected)
            continue
        nearby = [line for line in detected if abs(line - expected) <= 8]
        selected.append(min(nearby, key=lambda line: abs(line - expected)) if nearby else expected)
    selected = sorted(selected)
    selected[0] = max(0, selected[0])
    selected[-1] = min(width, selected[-1])
    return selected


def _fallback_column_lines(width: int) -> list[int]:
    return [round(width * ratio) for ratio in COLUMN_BOUNDARY_RATIOS]


def _fallback_row_lines(height: int) -> list[int]:
    estimated_rows = 12
    return [round(height * index / estimated_rows) for index in range(estimated_rows + 1)]


def _row_bands(lines: list[int], height: int) -> list[tuple[int, int]]:
    lines = [line for line in lines if 0 <= line <= height]
    if not lines or lines[0] > 8:
        lines.insert(0, 0)
    if height - lines[-1] > 8:
        lines.append(height)
    bands = []
    for top, bottom in zip(lines, lines[1:], strict=False):
        if bottom - top >= 22:
            bands.append((max(0, top), min(height, bottom)))
    return bands


def _data_row_bands(lines: list[int], height: int) -> list[tuple[int, int]]:
    lines = sorted({line for line in lines if 0 <= line <= height})
    if len(lines) < 3:
        return _row_bands(lines, height)

    header_bottom_candidates = [line for line in lines if line >= height * 0.15]
    header_bottom = header_bottom_candidates[0] if header_bottom_candidates else lines[1]
    data_lines = [line for line in lines if line >= header_bottom]
    if len(data_lines) < 2:
        return []

    small_bands = [(top, bottom) for top, bottom in zip(data_lines, data_lines[1:], strict=False) if bottom - top >= 12]
    if not small_bands:
        return []

    heights = [bottom - top for top, bottom in small_bands]
    if len(small_bands) >= 12 and mean(heights) < 38:
        grouped: list[tuple[int, int]] = []
        index = 0
        if len(small_bands) >= 2 and small_bands[1][1] - small_bands[0][0] <= 58:
            grouped.append((small_bands[0][0], small_bands[1][1]))
            index = 2
        while index < len(small_bands):
            chunk = small_bands[index : index + 3]
            if len(chunk) < 2:
                break
            grouped.append((chunk[0][0], chunk[-1][1]))
            index += 3
        return grouped

    return small_bands


def _save_cell_box_debug(image: np.ndarray, columns: list[tuple[int, int]], rows: list[tuple[int, int]], debug_dir: Path) -> None:
    overlay = image.copy()
    for row_index, (top, bottom) in enumerate(rows, start=1):
        for column_index, (left, right) in enumerate(columns, start=1):
            cv2.rectangle(overlay, (left, top), (right, bottom), (0, 0, 255), 1)
            if column_index == 1:
                cv2.putText(overlay, str(row_index), (left + 3, max(top + 15, 15)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 0, 0), 1)
    cv2.imwrite(str(debug_dir / "cell_boxes.png"), overlay)


def _crop_cell(image: np.ndarray, left: int, top: int, right: int, bottom: int) -> np.ndarray:
    height, width = image.shape[:2]
    pad_x = max(3, round((right - left) * 0.06))
    pad_y = max(3, round((bottom - top) * 0.10))
    crop = image[max(0, top + pad_y) : min(height, bottom - pad_y), max(0, left + pad_x) : min(width, right - pad_x)]
    if crop.size == 0:
        crop = image[max(0, top) : min(height, bottom), max(0, left) : min(width, right)]
    return crop


def _ocr_cell(crop: np.ndarray) -> tuple[str | None, float]:
    prepared = _prepare_cell_for_ocr(crop)
    try:
        result = _paddle_ocr().ocr(prepared)
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=503, detail="PaddlePaddle is not installed. Install paddlepaddle for local OCR.") from exc
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"PaddleOCR failed: {exc}") from exc

    texts: list[str] = []
    confidences: list[float] = []
    for page in result or []:
        if isinstance(page, dict):
            page_texts = page.get("rec_texts") or []
            page_scores = page.get("rec_scores") or []
            for text, confidence in zip(page_texts, page_scores, strict=False):
                text = str(text).strip()
                if text:
                    texts.append(text)
                    confidences.append(float(confidence))
            continue

        for item in page or []:
            if len(item) < 2:
                continue
            text = str(item[1][0]).strip()
            confidence = float(item[1][1])
            if text:
                texts.append(text)
                confidences.append(confidence)

    if not texts:
        return None, 0.0
    return " ".join(texts), round(mean(confidences), 3)


def _prepare_cell_for_ocr(crop: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, h=10)

    prepared = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        15,
        8,
    )

    height, width = prepared.shape[:2]
    ink = cv2.bitwise_not(prepared)
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(width // 2, 20), 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(height // 2, 12)))
    horizontal_lines = cv2.dilate(cv2.erode(ink, horizontal_kernel, iterations=1), horizontal_kernel, iterations=1)
    vertical_lines = cv2.dilate(cv2.erode(ink, vertical_kernel, iterations=1), vertical_kernel, iterations=1)
    grid_lines = cv2.dilate(cv2.bitwise_or(horizontal_lines, vertical_lines), np.ones((2, 2), np.uint8), iterations=1)
    prepared[grid_lines > 0] = 255

    prepared = cv2.copyMakeBorder(prepared, 20, 20, 20, 20, cv2.BORDER_CONSTANT, value=255)
    scale = 5 if max(prepared.shape[:2]) < 700 else 2
    prepared = cv2.resize(prepared, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    return cv2.cvtColor(prepared, cv2.COLOR_GRAY2BGR)


def _validate_row(row: dict) -> dict:
    uncertain_fields = set(row.get("uncertain_fields") or [])
    if row.get("s_no") and not re.search(r"\d+", str(row["s_no"])):
        uncertain_fields.add("s_no")
    if row.get("age_sex") and not re.search(r"\d{1,3}\s*/?\s*[mMfF]", str(row["age_sex"])):
        uncertain_fields.add("age_sex")
    if row.get("date_of_procedure") and not re.search(r"\d{1,2}[/.-]\d{1,2}[/.-]\d{2,4}", str(row["date_of_procedure"])):
        uncertain_fields.add("date_of_procedure")
    if row.get("ot") and not re.search(r"\d|ot", str(row["ot"]), flags=re.IGNORECASE):
        uncertain_fields.add("ot")
    row["uncertain_fields"] = [field for field in COLUMN_KEYS if field in uncertain_fields]
    return row


def _debug_dir(record_dir: Path) -> Path:
    debug_dir = record_dir / "debug"
    debug_dir.mkdir(parents=True, exist_ok=True)
    return debug_dir


@lru_cache(maxsize=1)
def _paddle_ocr():
    try:
        from paddleocr import PaddleOCR
    except Exception as exc:
        raise HTTPException(status_code=503, detail="PaddleOCR is not installed") from exc

    try:
        return PaddleOCR(
            lang="en",
            use_doc_orientation_classify=False,
            use_doc_unwarping=False,
            use_textline_orientation=False,
            text_det_limit_side_len=64,
        )
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Unable to initialize PaddleOCR: {exc}") from exc


def _normalize_payload(payload: dict) -> dict:
    rows = payload.get("rows")
    if not isinstance(rows, list):
        raise HTTPException(status_code=422, detail="Payload rows must be a list")

    normalized_rows = []
    for row in rows:
        source = row if isinstance(row, dict) else {}
        uncertain_fields = [field for field in source.get("uncertain_fields", []) if field in COLUMN_KEYS]
        normalized_rows.append(
            {
                **{column: _nullable_cell(source.get(column)) for column in COLUMN_KEYS},
                "confidence": _confidence(source.get("confidence")),
                "uncertain_fields": uncertain_fields,
            }
        )

    return {
        "document_type": DOCUMENT_TYPE,
        "columns": COLUMN_KEYS,
        "rows": normalized_rows,
    }


def _nullable_cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _confidence(value: object) -> float:
    try:
        return max(0.0, min(1.0, float(value)))
    except (TypeError, ValueError):
        return 0.0


def _record_dir(record_id: str) -> Path:
    return Path(settings.upload_dir) / "ot_register" / record_id


def _existing_record_dir(record_id: str) -> Path:
    if not record_id or Path(record_id).name != record_id:
        raise HTTPException(status_code=400, detail="Invalid record id")
    record_dir = _record_dir(record_id)
    if not record_dir.exists():
        raise HTTPException(status_code=404, detail="Record not found")
    return record_dir


def _active_json_path(record_dir: Path) -> Path:
    reviewed = record_dir / "reviewed.json"
    extracted = record_dir / "extracted.json"
    if reviewed.exists():
        return reviewed
    if extracted.exists():
        return extracted
    raise HTTPException(status_code=404, detail="No extracted data found for this record")


def _first_existing(record_dir: Path, pattern: str) -> str:
    return str(next(record_dir.glob(pattern), ""))


def _read_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail=f"Invalid stored JSON: {path.name}") from exc


def _write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _record_response(record_id: str, data: dict, image_path: str, processed_path: str) -> dict:
    return {
        "id": record_id,
        "imagePath": image_path,
        "processedImagePath": processed_path,
        "data": data,
    }
